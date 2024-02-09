from unittest.mock import Mock, patch
from openpyxl import load_workbook
import pandas as pd
import pytest
from io import BytesIO
from src.books import Book, BookColumns, EmptyColumns
from src.one_lakh_plus_transactions import TransactionAbove1L
from src.report import Report
from src.filingmonth import FilingMonth


mock_engine = Mock(return_value="mocked_engine")
mock_data = {
    "sql": "SELECT * FROM test_table",
    "details": {
        'PAN': '1234567890',
        'office_name': 'test_office'
    }
}
mock_transactions_data = {
    "Transaction ID": ["T1", "T2", "T3", "T4"],
    "Bill Receiveable Person": ["ABC Inn", "Sajha ban", "Robert Amdocs", "John Doe"],
    "Vat Pan No": [123, 234, 345, 567],
    "Grand Total": [52_987, 2_000, 5_000, 2_00_000],
    "Taxable Amount": [46891.15, 1769.91, 4424.78, 176991.15],
    "Status": ["001-03", "001-01", "001-03", "001-01"],
    "Modify Type": ["Cancel", "", "Cancel", ""],
    # ...other columns not required for this testing.
}


@pytest.fixture
def mock_db_transactions():
    return pd.DataFrame(mock_transactions_data)


@pytest.fixture
def mock_raw_transactions():
    """
    mock raw transactions filtering cancelled transactions.
    """
    df = pd.DataFrame(mock_transactions_data)

    # Drop rows where Status is "001-03"
    filtered_df = df[df['Status'] != "001-03"]
    return filtered_df


@pytest.fixture
def mock_transactions(mock_raw_transactions):
    """
    represents transactions that are output of process_transactions.
    """
    # hardcoded
    index = 1
    drop_columns = ["Status", "Modify Type"]
    mock_raw_transactions.insert(index, index, None)
    mock_transactions = mock_raw_transactions.drop(drop_columns, axis=1)
    return mock_transactions


@pytest.fixture(scope="module")
def monkeymodule():
    with pytest.MonkeyPatch.context() as mp:
        yield mp


@pytest.fixture(scope="function")
def mock_sql_engine(monkeymodule):
    mock_sql_engine = mock_engine
    monkeymodule.setattr('src.report.get_sql_engine', mock_sql_engine)


@pytest.fixture(scope="function")
def mock_get_data(monkeymodule):
    monkeymodule.setattr('src.report.get_data', lambda x: mock_data[x])


@pytest.fixture
def saved_mock_get_template_buffer():
    with open('tests/io_files/sales.xlsx', "rb") as fh:
        template_buffer = BytesIO(fh.read())
    with patch('src.report.Report.get_template_buffer', return_value=template_buffer):
        yield template_buffer

# Arrange a test directory


@pytest.fixture
def work_dir(tmp_path):
    test_work_dir = tmp_path / "work_dir"
    test_work_dir.mkdir()
    return test_work_dir


@pytest.fixture
def filing_month_instance():
    return FilingMonth(2080, 7)


@pytest.fixture
def book_instance():
    return Book(
        id=1,
        name='sales',
        sheet='Nepali SB',
        symbol='S',
        columns=BookColumns([
            "Transaction ID",
            "Bill Receiveable Person",
            "Vat Pan No",
            "Grand Total",
            "Taxable Amount",
        ]),
        endpoint='/book',
        emptycols=EmptyColumns([1]),
    )


@pytest.fixture
def test_report(filing_month_instance, book_instance, work_dir, mock_sql_engine, mock_get_data):
    return Report(book_instance, filing_month_instance, work_dir)


def test_report_init(test_report, book_instance, work_dir, filing_month_instance):
    report_instance = test_report

    # Assert
    assert report_instance.book == book_instance
    assert report_instance.filing_month == filing_month_instance
    assert report_instance.date_range == filing_month_instance.get_AD_date_range()
    assert report_instance.save_filepath == work_dir.joinpath(
        f"{book_instance.name} - {report_instance.filing_month.nepali_month_name()}.xlsx")
    assert report_instance.cancelled_transactions == []
    assert isinstance(report_instance.buffer, BytesIO)
    assert report_instance._raw_transactions is None
    assert report_instance._transactions is None


def test_remove_cancelled_transactions_removes_correctly(test_report, mock_db_transactions, mock_raw_transactions):
    # Create sample DataFrame with cancelled and non-cancelled transactions
    df = pd.DataFrame(mock_db_transactions)

    # Call the function to remove cancelled transactions
    filtered_df = test_report.remove_cancelled_transactions(df.copy())

    assert isinstance(filtered_df, pd.DataFrame)

    # Assert that only non-cancelled transactions remain
    assert filtered_df.shape[0] == 2
    assert sorted(filtered_df["Transaction ID"].tolist()) == ["T2", "T4"]
    pd.testing.assert_frame_equal(
        filtered_df.reset_index(drop=True),
        mock_raw_transactions.reset_index(drop=True)
    )


# @patch('src.report.get_sql_engine')
@patch('src.report.pd.read_sql')
def test_query_db(
    mock_read_sql,
    test_report,
    book_instance,
    filing_month_instance,
    mock_db_transactions,
    mock_raw_transactions,
):

    mock_read_sql.return_value = mock_db_transactions

    # Call the query_db function
    raw_transactions = test_report.query_db()

    date_range = filing_month_instance.get_AD_date_range()

    # Assert that the correct parameters were passed to read_sql
    mock_read_sql.assert_called_once_with(
        mock_data['sql'],
        mock_engine.return_value,
        params=(
            book_instance.id,
            date_range.start,
            date_range.end,
        )
    )

    # Assert that the returned DataFrame is the expected one
    pd.testing.assert_frame_equal(
        mock_raw_transactions.reset_index(drop=True),
        raw_transactions.reset_index(drop=True)
    )


def test_process_transactions(test_report, mock_raw_transactions, mock_transactions):

    test_report._raw_transactions = mock_raw_transactions

    # Test
    result = test_report.process_transactions()

    # Assert
    assert isinstance(result, pd.DataFrame)
    pd.testing.assert_frame_equal(
        result.reset_index(drop=False),
        mock_transactions.reset_index(drop=False)
    )


def test_fill_report_details(test_report, saved_mock_get_template_buffer):
    # Test
    test_report.fill_report_details()
    workbook = load_workbook(test_report.buffer)
    sheet = workbook.active

    # Assert
    assert sheet['A4'].value == 'करदाता दर्ता नं (PAN) : 1234567890        करदाताको नाम: test_office         साल: 2080    कर अवधि: Kartik'


def test_populate_report_buffer(test_report, mock_transactions, saved_mock_get_template_buffer, monkeypatch):
    # Mocking fill_report_details for simplicity
    monkeypatch.setattr(
        "src.report.Report.fill_report_details", lambda x: None)
    test_report.buffer = test_report.get_template_buffer()

    test_report._transactions = mock_transactions
    # test_report.fill_report_details()

    # Test
    test_report.populate_report_buffer()
    workbook = load_workbook(test_report.buffer)
    sheet = workbook.active

    # Assert
    assert sheet['A7'].value == 'T2'
    assert sheet['B7'].value == None
    assert sheet['C7'].value == 'Sajha ban'


def test_save(test_report, monkeypatch):
    test_report.buffer = BytesIO(b'This is a test')
    monkeypatch.setattr(
        'src.report.Report.populate_report_buffer', lambda _: None)

    with patch('src.report.write_bytes_to_disk', Mock()) as mock_writer:
        test_report.save()

    # Assert that the write_bytes_to_disk function was called with the correct arguments
    mock_writer.assert_called_once_with(
        test_report.buffer, test_report.save_filepath)


def test_get_transactions_above_1L(test_report, mock_transactions):
    # Set the raw transactions data for the test report
    test_report._transactions = mock_transactions

    # # Mock the process_transactions method since it's used in get_transactions_above_1L
    # with patch.object(test_report, 'process_transactions', return_value=sample_raw_transactions):

    # Call the method under test
    result = test_report.get_transactions_above_1L()

    # Define the expected result based on your sample data
    expected_result = [
        TransactionAbove1L(pan_no=567, bill_receiveable_person='John Doe',
                           trade_name_type='E', transaction_type='S', taxable_amount=176991.15, exempted_amount=0),
    ]

    # Perform the assertion
    assert result == expected_result
